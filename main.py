import openpyxl as oe

excel_document = oe.load_workbook("Glyph Union.xlsx")


#print(excel_document.get_sheet_names())

hoja = excel_document["Hoja1"]

#print(hoja["A1"].value)





# for image in hoja._images:
#         print(image.path)

for x in range(2, 100):
    if hoja.cell(row=x, column=1).value is not None:
        celda = hoja.cell(row=x, column=1)
 #       print(celda.value.find("="))
        equal = celda.value.find("=")
        plus=celda.value.find("+")
        par = celda.value.find("(")
        celda.value= celda.value[equal+1:par]
        #print(celda.value[equal+1:par].strip(),x)
        #first = celda.value[:plus]
        #second= celda.value[plus+1: equal]
        #print(f"el primer item es {first} y el segundo {second}")
        #hoja[f"G{x}"] = first
        #hoja[f"H{x}"] = second

# for x in hoja.columns:
#         for y in hoja.rows:

#                 print(hoja.cell(row=x, column=y).hyperlink.target)

#     print(celda.value)
#     if celda.value is not None:
#             celda.value = celda.value.replace(" - ", '')
#     else:
#         continue


excel_document.save("Glyph Union.xlsx")
